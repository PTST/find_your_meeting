import numpy as np
import json
import sys
from pathfinding.core.diagonal_movement import DiagonalMovement
from pathfinding.core.grid import Grid
from pathfinding.finder.a_star import AStarFinder
import openpyxl
import os
import time
import concurrent.futures


'''
Import variables from external files
'''
basement_list = list(np.load("resources/Basement.npy"))
ground_floor_list = list(np.load("resources/GroundFloor.npy"))
first_floor_list = list(np.load("resources/SecondFloor.npy"))
tower_list = list(np.load("resources/Tower.npy"))
with open("resources/locations.json", "r") as f:
    jsonfile = f.read()
locations = json.loads(jsonfile)
with open("resources/up_down.json", "r") as f:
    jsonfile = f.read()
up_down = json.loads(jsonfile)

def select_start_end(loc_dict):
    while True:
        while True:
            try:
                start_point = input("\nStart: ").upper()
                test = loc_dict[start_point][0]
                break
            except KeyError:
                print("\nNot a valid starting point. Try again")

        while True:
            try:
                end_point = input("\nEnd: ").upper()
                test2 = loc_dict[end_point][0]
                break
            except KeyError:
                print("\nNot a valid endpoint. Try again")
        if test == test2:
            print("Starting point and endpoint is the same")
        else:
            break
    return start_point, end_point

def find_path(location_dict, start_point, end_point, grid):
    start = grid.node(location_dict[start_point][0], location_dict[start_point][1])
    end = grid.node(location_dict[end_point][0], location_dict[end_point][1])
    finder = AStarFinder(diagonal_movement=DiagonalMovement.never)
    path, runs = finder.find_path(start, end, grid)

    print("operations:", runs, "path length:", len(path))
    return path

def load_excel():
    operating_system = sys.platform
    if operating_system == "win32":
        wb = openpyxl.load_workbook("M:\\GitHub\\find_your_meeting\\floorPlan.xlsm")
    elif operating_system == "darwin":
        wb = openpyxl.load_workbook("/Users/PTST/Dev/find_your_meeting/floorPlan.xlsm")
    else:
        raise ValueError("Unknown operating system")
    return wb

def save_to_excel(wb, path, floors):
    operating_system = sys.platform
    for i in range(len(path)):
        print(floors[i])
        print(type(floors[i]))
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[int(floors[i])])
        for item in path[i]:
            sheet.cell(row=(item[1]+1), column=(item[0]+1)).value = "x"
            sheet.cell(row=(item[1]+1), column=(item[0]+1)).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
            sheet.cell(row=(item[1]+1), column=(item[0]+1)).fill = openpyxl.styles.PatternFill(fgColor=openpyxl.styles.colors.RED, fill_type = "solid")
    if operating_system == "win32":
        wb.save("M:\\GitHub\\find_your_meeting\path.xlsx")
        os.startfile("M:\\GitHub\\find_your_meeting\path.xlsx")
    elif operating_system == "darwin":
        wb.save("/Users/PTST/Dev/find_your_meeting/path.xlsx")
        os.system("open /Users/PTST/Dev/find_your_meeting/path.xlsx")
    else:
        raise ValueError("Unknown operating system")

def multiple_floors(counter, floor_no, going_up):
    if floor_no == 0:
        floor = basement_list
    elif floor_no == 1:
        floor = ground_floor_list
    elif floor_no == 2:
        floor = first_floor_list
    elif floor_no == 3:
        floor = tower_list
    else:
        raise ValueError("WTF")
    
    if counter == 0:
        if going_up:
            direction = "UP"
        else:
            direction = "DOWN"
    else:
        if going_up:
            direction = "DOWN"
        else:
            direction = "UP"
    executor = concurrent.futures.ProcessPoolExecutor(20)
    futures = [executor.submit(get_floor_paths, key, direction, floor_no, floor) for key, value in up_down.items()]
    concurrent.futures.wait(futures)
        #get_floor_paths(key, direction, floor_no, floor)

def get_floor_paths(a, direction, floor_no, floor):
    if a.endswith(direction) and a.startswith(str(floor_no)):
        search_grid = Grid(matrix=floor)
        if counter == 0:
            found_path = find_path(locations, start_loc, a, search_grid)
            if len(found_path) != 0:
                paths_start[a] = found_path
        elif counter == 1:
            found_path = find_path(locations, a, end_loc, search_grid)
            if len(found_path) != 0:
                paths_end[a] = found_path

start_time = time.time()
#start_loc, end_loc = select_start_end(locations)
start_loc, end_loc = "H2.06", "H1.25"
start_floor, end_floor = start_loc[1], end_loc[1]

if start_floor != end_floor:
    paths_start = {}
    paths_end = {}
    upwards = (start_floor < end_floor)
    print(upwards)
    for counter, floor_id in enumerate([int(start_floor), int(end_floor)]):
        multiple_floors(counter, floor_id, upwards)
        

best_value = sys.maxsize
best_path = []
for key, value in paths_start.items():
    for key2, value2 in paths_end.items():
        if key[2:3] == key2[2:3]:
            total = len(value) + len(value2)
            if total < best_value:
                best_value = total
                best_path = []
                best_path.append(value)
                best_path.append(value2)

print("--- %s seconds ---" % (time.time() - start_time))

#workbook = load_excel()
#save_to_excel(workbook, best_path, [start_floor, end_floor])


        

