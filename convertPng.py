from PIL import Image

import openpyxl
wb = openpyxl.load_workbook('test.xlsx')
sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

im = Image.open("basement.png") #Can be many different formats.
pix = im.load()
img_x, img_y = tuple(map(int, im.size))

white = (255,255,255,255)
print(pix[5,5] != white)
for y in range(0, img_y):
    for x in range(0, img_x):
        if pix[x,y] != white:
        	sheet.cell(row=int((y+1)), column=int((x+1))).value = "#"
            # sheet.cell(row=int((y+1)), column=int((x+1))).fill = openpyxl.styles.PatternFill(fgColor="000000", fill_type = "solid")
wb.save('basement2.xlsx')
