import openpyxl
from pathfinding.core.diagonal_movement import DiagonalMovement
from pathfinding.core.grid import Grid
from pathfinding.finder.a_star import AStarFinder
import sys
import os
import copy

"""
Loads excel file into memory, extracts the first sheet as an object, and returns the numerical value of 'lastColumn'
"""
def load_excel(lastColumn):
    operating_system = sys.platform
    if operating_system == "win32":
        wb = openpyxl.load_workbook("M:\\GitHub\\find_your_meeting\\groundFloor.xlsm")
    elif operating_system == "darwin":
        wb = openpyxl.load_workbook("/Users/PTST/Dev/find_your_meeting/groundFloor.xlsm")
    else:
        raise ValueError("Unknown operating system")
    return wb, wb.get_sheet_by_name(wb.get_sheet_names()[0]), openpyxl.utils.column_index_from_string(lastColumn)

'''
Converts excel sheet to a square grid with the dimension last_cell x last_cell
'''
def excel_to_list(sheet, last_cell):
    master_list = []
    location_dict = {}
    up_down_dict = {}
    for y in range(1, last_cell+1):
        x_list = []
        for x in range(1, last_cell+1):
            cell_value = sheet.cell(row=y, column=x).value
            if cell_value == ".":
                x_list.append(0)
            elif cell_value is not None:
                if (cell_value.upper().startswith("UP")) or (cell_value.upper().startswith("DOWN")):
                    up_down_dict[str(cell_value).upper()] = [x-1, y-1]
                    location_dict[str(cell_value).upper()] = [x-1, y-1]
                else:
                    location_dict[str(cell_value).upper()] = [x-1, y-1]
                x_list.append(0)
            else:
                x_list.append(1)
        master_list.append(x_list)
    
    return master_list, location_dict, up_down_dict

'''
Prints available location from dict
'''
def print_locations(location_dict):
    print("\nPossible locations:")
    for key, value in location_dict.items():
        if not (key.startswith("UP")) or (key.startswith("DOWN")):
            print (key, value)
'''
Get start and endpoint coordinates
'''
def get_coordinates(location_dict):
    while True:
        while True:
            try:
                start_point = input("\nStart: ").upper()
                test = location_dict[start_point][0]
                break
            except KeyError:
                print("\nNot a valid starting point. Try again")
                print_locations(location_dict)

        while True:
            try:
                end_point = input("\nEnd: ").upper()
                test2 = location_dict[end_point][0]
                break
            except KeyError:
                print("\nNot a valid endpoint. Try again")
                print_locations(location_dict)
        if test == test2:
            print("Starting point and endpoint is the same")
        else:
            break
    return start_point, end_point

'''
Finds optimal path between locations via a* algorithm
'''
def find_path(location_dict, start_point, end_point, grid):
    start = grid.node(location_dict[start_point][0], location_dict[start_point][1])
    end = grid.node(location_dict[end_point][0], location_dict[end_point][1])
    finder = AStarFinder(diagonal_movement=DiagonalMovement.never)
    path, runs = finder.find_path(start, end, grid)

    print("operations:", runs, "path length:", len(path))
    return path

'''
saves excel workbook to file with optimal path highlighted
'''
def save_to_excel(wb, path, sheet, open_excel=True):
    operating_system = sys.platform
    for item in path:
        sheet.cell(row=(item[1]+1), column=(item[0]+1)).value = "x"
        sheet.cell(row=(item[1]+1), column=(item[0]+1)).font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
        sheet.cell(row=(item[1]+1), column=(item[0]+1)).fill = openpyxl.styles.PatternFill(fgColor=openpyxl.styles.colors.RED, fill_type = "solid")
    if operating_system == "win32":
        wb.save("M:\\GitHub\\find_your_meeting\path.xlsx")
        if open_excel:
            os.startfile("M:\\GitHub\\find_your_meeting\path.xlsx")
    elif operating_system == "darwin":
        wb.save("/Users/PTST/Dev/find_your_meeting/path.xlsx")
        if open_excel:
            os.system("open /Users/PTST/Dev/find_your_meeting/path.xlsx")
    else:
        raise ValueError("Unknown operating system")


'''
Execution of functions
'''
#last_column=input("Last column: ").upper()
last_column = "OQ"
workbook, excel_sheet, max_cell = load_excel(last_column)
orig_workbook = copy.copy(workbook)
master_list, possible_locations, up_down_locations = excel_to_list(excel_sheet, max_cell)
search_grid = Grid(matrix=master_list)


possible_locations["H2.08"] = [200,300]


print_locations(possible_locations)
start_node, end_node = get_coordinates(possible_locations)
path_dict_start = {}
if start_node[1:2] != end_node[1:2]:
    if start_node[1:2] < end_node[1:2]:
        direction = "UP"
    else:
        direction = "DOWN"
    for key, value in up_down_locations.items():
        if key.startswith(direction):
            found_path = find_path(possible_locations, start_node, key, search_grid)
            path_dict_start[key] = found_path
            search_grid = Grid(matrix=master_list)
else:
    found_path = find_path(possible_locations, start_node, end_node, search_grid)

short_key = min(path_dict_start, key=path_dict_start.get)
print(short_key + " - " + str(len(path_dict_start[short_key])))
#save_to_excel(workbook, found_path, excel_sheet, False)